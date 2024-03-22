from openpyxl import Workbook
from openpyxl.styles import Font
import mysql.connector

connection = mysql.connector.connect(
    host="",
    user="",
    password="",
    database=""
)

wb = Workbook()

ws_compteClient = wb.active
ws_compteClient.title = "Compte Client"
ws_compteClient['A1'] = "user_nicename"
ws_compteClient['B1'] = "ID"
ws_compteClient['C1'] = "_billing_email"
ws_compteClient['D1'] = "_billing_fist_name"
ws_compteClient['E1'] = "_billing_last_name"
ws_compteClient['F1'] = "_billing_adresses_1"
ws_compteClient['G1'] = "_billing_city"
ws_compteClient['H1'] = "_billing_postcode"
ws_compteClient['I1'] = "_billing_phone"
ws_compteClient['J1'] = "date_of_birth"

for cell in ws_compteClient[1]:
    cell.font = Font(color="FF0000", bold=True)

cursor = connection.cursor()
query_client = """
SELECT
    u.user_nicename,
    u.ID,
    um1.meta_value AS billing_email,
    um2.meta_value AS billing_first_name,
    um3.meta_value AS billing_last_name,
    um4.meta_value AS billing_address_1,
    um5.meta_value AS billing_city,
    um6.meta_value AS billing_postcode,
    um7.meta_value AS billing_phone,
    um8.meta_value AS date_of_birth
FROM
    wp_users u
LEFT JOIN
    wp_usermeta um1 ON u.ID = um1.user_id AND um1.meta_key = 'billing_email'
LEFT JOIN
    wp_usermeta um2 ON u.ID = um2.user_id AND um2.meta_key = 'billing_first_name'
LEFT JOIN
    wp_usermeta um3 ON u.ID = um3.user_id AND um3.meta_key = 'billing_last_name'
LEFT JOIN
    wp_usermeta um4 ON u.ID = um4.user_id AND um4.meta_key = 'billing_address_1'
LEFT JOIN
    wp_usermeta um5 ON u.ID = um5.user_id AND um5.meta_key = 'billing_city'
LEFT JOIN
    wp_usermeta um6 ON u.ID = um6.user_id AND um6.meta_key = 'billing_postcode'
LEFT JOIN
    wp_usermeta um7 ON u.ID = um7.user_id AND um7.meta_key = 'billing_phone'
LEFT JOIN
    wp_usermeta um8 ON u.ID = um8.user_id AND um8.meta_key = 'bcn_birthday_date'
ORDER BY `u`.`ID` ASC
"""


cursor.execute(query_client)

for row_index, row_data in enumerate(cursor.fetchall(), start=2):
    for col_index, cell_value in enumerate(row_data, start=1):
        ws_compteClient.cell(row=row_index, column=col_index, value=cell_value)

cursor.close()

for column_cells in ws_compteClient.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws_compteClient.column_dimensions[column_cells[0].column_letter].width = length + 2


ws_commande = wb.create_sheet(title="Commande")

ws_commande['A1'] = "post_id"
ws_commande['B1'] = "post_date"
ws_commande['C1'] = "salername"
ws_commande['D1'] = "post_status"
ws_commande['E1'] = "order_total"
ws_commande['F1'] = "order_shipping"
ws_commande['G1'] = "payment_method"
ws_commande['H1'] = "customer_user"
ws_commande['I1'] = "remises"
ws_commande['J1'] = "numéro_facture"
ws_commande['K1'] = "coupon_code"

for cell in ws_commande[1]:
    cell.font = Font(color="FF0000", bold=True)


cursor = connection.cursor()

query_commande = """
SELECT 
    wp_postmeta.post_id,
    wp_posts.post_date,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE 'seller_name' THEN wp_postmeta.meta_value ELSE NULL END) AS seller_name,
    wp_posts.post_status,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_order_total' THEN wp_postmeta.meta_value ELSE NULL END) AS order_total,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_order_shipping' THEN wp_postmeta.meta_value ELSE NULL END) AS order_shipping,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_payment_method' THEN wp_postmeta.meta_value ELSE NULL END) AS payment_method,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_customer_user' THEN wp_postmeta.meta_value ELSE NULL END) AS customer_user,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_cart_discount' THEN wp_postmeta.meta_value ELSE NULL END) AS remises,
    MAX(CASE WHEN wp_postmeta.meta_key LIKE '_wcpdf_invoice_number' THEN wp_postmeta.meta_value ELSE NULL END) AS numéro_facture,
    GROUP_CONCAT(DISTINCT 
        CASE 
            WHEN wp_woocommerce_order_items.order_item_name LIKE '%cadeau commande%' THEN 'cadeau commande' 
            WHEN wp_woocommerce_order_items.order_item_name LIKE '%carte cadeau%' THEN 'carte cadeau'
            ELSE 'NULL'
        END SEPARATOR ', ') AS coupon_code
FROM wp_postmeta 
JOIN wp_posts ON wp_postmeta.post_id = wp_posts.ID
LEFT JOIN wp_woocommerce_order_items ON wp_woocommerce_order_items.order_id = wp_postmeta.post_id
WHERE 
    (wp_postmeta.meta_key LIKE '_order_total' 
    OR wp_postmeta.meta_key LIKE '_payment_method'
    OR wp_postmeta.meta_key LIKE '_customer_user'
    OR wp_postmeta.meta_key LIKE '_order_shipping'
    OR wp_postmeta.meta_key LIKE '_customer_user'
    OR wp_postmeta.meta_key LIKE 'seller_name'
    OR wp_postmeta.meta_key LIKE '_cart_discount' 
    OR wp_postmeta.meta_key LIKE '_wcpdf_invoice_number')
GROUP BY wp_postmeta.post_id, wp_posts.post_date, wp_posts.post_status
ORDER BY wp_postmeta.post_id
"""

cursor.execute(query_commande)

for row_data in cursor.fetchall():
    ws_commande.append(row_data)



for column_cells in ws_commande.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws_commande.column_dimensions[column_cells[0].column_letter].width = length + 2



ws_produit = wb.create_sheet(title="Produits")

ws_produit['A1'] = "post_id"
ws_produit['B1'] = "item_name"
ws_produit['C1'] = "quantity"

for cell in ws_produit[1]:
    cell.font = Font(color="FF0000", bold=True)

cursor = connection.cursor()

query_produit = """
SELECT 
    orders.ID AS post_id,
    products.post_title AS item_name,
    quantity.meta_value AS quantity
FROM 
    wp_posts AS orders
JOIN 
    wp_woocommerce_order_items AS order_items ON orders.ID = order_items.order_id
JOIN 
    wp_woocommerce_order_itemmeta AS product_id_meta ON order_items.order_item_id = product_id_meta.order_item_id AND product_id_meta.meta_key = '_product_id'
JOIN 
    wp_posts AS products ON product_id_meta.meta_value = products.ID
JOIN 
    wp_woocommerce_order_itemmeta AS quantity ON order_items.order_item_id = quantity.order_item_id AND quantity.meta_key = '_qty'
WHERE 
    orders.post_type = 'shop_order'
GROUP BY
    orders.ID, products.post_title, quantity.meta_value
ORDER BY 
    orders.ID
"""


cursor.execute(query_produit)

for row_data in cursor.fetchall():
    ws_produit.append(row_data)

cursor.close()

for column_cells in ws_produit.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws_produit.column_dimensions[column_cells[0].column_letter].width = length + 2



wb.save("salesforce.xlsx")
print("Fichier Excel créé avec succès.")
